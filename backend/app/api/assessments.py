"""
检查数据管理 API（研究级）
"""
from typing import List, Optional, Any
from datetime import datetime
import json
import io
import csv
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.core.database import get_db
from app.core.permissions import get_study_member_or_403, get_study_id_from_visit
from app.models.tables import AssessmentData, Visit, Subject, Study, StudyMember, User
from app.schemas.schemas import AssessmentDataResponse, AssessmentDataUpdate
from app.api.auth import get_current_active_user

router = APIRouter()


class ManualAssessmentCreate(BaseModel):
    """手动录入检查数据请求"""
    visit_id: int
    assessment_type: str
    extracted_data: dict[str, Any]
    file_id: Optional[int] = None
    sample_time: Optional[datetime] = None


@router.post("/manual", response_model=AssessmentDataResponse)
async def create_manual_assessment(
    assessment: ManualAssessmentCreate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """手动录入检查数据"""
    # 检查随访记录是否存在
    visit = db.query(Visit).filter(Visit.id == assessment.visit_id).first()
    if not visit:
        raise HTTPException(status_code=404, detail="随访记录不存在")

    # 获取研究 ID 并检查权限
    study_id = visit.subject.study_id
    _get_study_member_or_403(study_id, current_user, db)

    db_assessment = AssessmentData(
        visit_id=assessment.visit_id,
        file_id=assessment.file_id,
        assessment_type=assessment.assessment_type,
        extracted_data=assessment.extracted_data,
        sample_time=assessment.sample_time,
        is_verified=True,  # 手动录入的数据默认已验证
        verified_at=datetime.utcnow(),
        verified_by=current_user.id,
    )
    db.add(db_assessment)
    db.commit()
    db.refresh(db_assessment)

    # 构建响应，避免返回数据库模型
    return AssessmentDataResponse(
        id=db_assessment.id,
        visit_id=db_assessment.visit_id,
        file_id=db_assessment.file_id,
        assessment_type=db_assessment.assessment_type,
        extracted_data=db_assessment.extracted_data,
        is_verified=db_assessment.is_verified,
        sample_time=db_assessment.sample_time,
        verified_at=db_assessment.verified_at,
        verified_by=db_assessment.verified_by,
        created_at=db_assessment.created_at,
        updated_at=db_assessment.updated_at,
    )


@router.get("/")
async def list_assessments(
    study_id: int = Query(..., description="研究 ID"),
    visit_id: Optional[int] = Query(None),
    assessment_type: Optional[str] = Query(None),
    is_verified: Optional[bool] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """获取检查数据列表"""
    # 检查权限
    _get_study_member_or_403(study_id, current_user, db)

    query = db.query(AssessmentData, Visit, Subject).join(
        Visit, AssessmentData.visit_id == Visit.id
    ).join(
        Subject, Visit.subject_id == Subject.id
    ).filter(Subject.study_id == study_id)

    if visit_id:
        query = query.filter(AssessmentData.visit_id == visit_id)
    if assessment_type:
        query = query.filter(AssessmentData.assessment_type == assessment_type)
    if is_verified is not None:
        query = query.filter(AssessmentData.is_verified == is_verified)

    results = query.offset(skip).limit(limit).all()

    # 返回带有受试者信息的数据
    data = []
    for assessment, visit, subject in results:
        item = {
            "id": assessment.id,
            "visit_id": assessment.visit_id,
            "file_id": assessment.file_id,
            "assessment_type": assessment.assessment_type,
            "extracted_data": assessment.extracted_data,
            "is_verified": assessment.is_verified,
            "sample_time": assessment.sample_time,
            "verified_at": assessment.verified_at,
            "verified_by": assessment.verified_by,
            "created_at": assessment.created_at,
            "updated_at": assessment.updated_at,
            # 额外添加受试者和随访信息
            "subject_code": subject.subject_code,
            "subject_name_pinyin": subject.name_pinyin,
            "visit_name": visit.visit_name,
            "visit_date": visit.visit_date.isoformat() if visit.visit_date else None,
        }
        data.append(item)

    return data


@router.get("/export")
async def export_assessments(
    format: str = Query("json"),
    assessment_type: Optional[str] = Query(None),
    is_verified: Optional[str] = Query(None),  # "true" or "false"
    data_format: Optional[str] = Query("wide", description="wide: 宽格式，long: 长格式"),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """导出评估数据

    参数:
        format: 导出文件格式 (json, csv, excel)
        assessment_type: 检测类型筛选
        is_verified: 是否已审核
        data_format: 数据格式 (wide: 宽格式-每行一个样本，long: 长格式-每行一个指标)
    """
    # 验证 format 参数
    if format not in ["json", "csv", "excel"]:
        raise HTTPException(status_code=400, detail="不支持的导出格式")

    if data_format not in ["wide", "long"]:
        raise HTTPException(status_code=400, detail="不支持的数据格式，仅支持 wide 或 long")

    query = db.query(AssessmentData, Visit, Subject).join(Visit, AssessmentData.visit_id == Visit.id).join(Subject, Visit.subject_id == Subject.id)

    if assessment_type:
        query = query.filter(AssessmentData.assessment_type == assessment_type)

    # 处理 is_verified 参数
    if is_verified is not None and is_verified != '':
        verified_bool = is_verified.lower() == 'true'
        query = query.filter(AssessmentData.is_verified == verified_bool)

    results = query.all()

    if data_format == "long":
        # 长格式：每行一个测量值
        # 列：样本编号，访视名称，访视日期，检测类型，指标名称，指标值，单位，是否审核，采样时间
        # 元数据字段（不作为生物标志物导出）
        metadata_fields = {"sample_time", "test_date", "created_at", "updated_at", "verified_at", "verified_by"}
        long_data = []
        for assessment, visit, subject in results:
            extracted = assessment.extracted_data or {}
            for field_name, value in extracted.items():
                # 跳过元数据字段
                if field_name in metadata_fields:
                    continue
                long_data.append({
                    "样本编号": subject.subject_code,
                    "访视名称": visit.visit_name,
                    "访视日期": visit.visit_date.strftime("%Y-%m-%d") if visit.visit_date else "",
                    "检测类型": assessment.assessment_type,
                    "指标名称": field_name,
                    "指标值": value if not isinstance(value, dict) else json.dumps(value, ensure_ascii=False),
                    "是否审核": "是" if assessment.is_verified else "否",
                    "采样时间": assessment.sample_time.strftime("%Y-%m-%d %H:%M") if assessment.sample_time else "",
                    "创建时间": assessment.created_at.strftime("%Y-%m-%d %H:%M") if assessment.created_at else "",
                })

        if format == "json":
            return Response(
                content=json.dumps(long_data, indent=2, ensure_ascii=False),
                media_type="application/json",
                headers={"Content-Disposition": f"attachment; filename=assessments_long_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"}
            )
        elif format == "csv":
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=["样本编号", "访视名称", "访视日期", "检测类型", "指标名称", "指标值", "是否审核", "采样时间", "创建时间"])
            writer.writeheader()
            writer.writerows(long_data)
            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=assessments_long_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
            )
        elif format == "excel":
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Assessments"
            headers = ["样本编号", "访视名称", "访视日期", "检测类型", "指标名称", "指标值", "是否审核", "采样时间", "创建时间"]
            ws.append(headers)
            for row in long_data:
                ws.append([row[h] for h in headers])
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=assessments_long_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )

    else:  # wide format
        # 宽格式：每行一个样本，动态列名为 检测类型 - 指标名
        # 列：样本编号，访视名称，访视日期，检测类型，各项指标...
        # 元数据字段（不作为生物标志物导出）
        metadata_fields = {"sample_time", "test_date", "created_at", "updated_at", "verified_at", "verified_by"}

        # 收集所有可能的指标字段（排除元数据）
        all_fields = set()
        for assessment, _, _ in results:
            if assessment.extracted_data:
                for key in assessment.extracted_data.keys():
                    if key not in metadata_fields:
                        all_fields.add(key)
        all_fields = sorted(all_fields)

        # 构建宽格式数据
        wide_data = []
        for assessment, visit, subject in results:
            row = {
                "样本编号": subject.subject_code,
                "访视名称": visit.visit_name,
                "访视日期": visit.visit_date.strftime("%Y-%m-%d") if visit.visit_date else "",
                "检测类型": assessment.assessment_type,
                "是否审核": "是" if assessment.is_verified else "否",
                "采样时间": assessment.sample_time.strftime("%Y-%m-%d %H:%M") if assessment.sample_time else "",
                "创建时间": assessment.created_at.strftime("%Y-%m-%d %H:%M") if assessment.created_at else "",
            }
            extracted = assessment.extracted_data or {}
            for field in all_fields:
                value = extracted.get(field, "")
                row[f"{assessment.assessment_type}_{field}"] = value if not isinstance(value, (dict, list)) else json.dumps(value, ensure_ascii=False)
            wide_data.append(row)

        if format == "json":
            return Response(
                content=json.dumps(wide_data, indent=2, ensure_ascii=False),
                media_type="application/json",
                headers={"Content-Disposition": f"attachment; filename=assessments_wide_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"}
            )
        elif format == "csv":
            # 动态生成列名（排除元数据字段）
            metadata_fields = {"sample_time", "test_date", "created_at", "updated_at", "verified_at", "verified_by"}
            field_cols = [f"{a.assessment_type}_{f}" for a, _, _ in results for f in (a.extracted_data or {}).keys() if f not in metadata_fields]
            field_cols = sorted(set(field_cols))
            all_cols = ["样本编号", "访视名称", "访视日期", "检测类型", "是否审核", "采样时间", "创建时间"] + field_cols

            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=all_cols)
            writer.writeheader()
            writer.writerows(wide_data)
            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=assessments_wide_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
            )
        elif format == "excel":
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            # 动态生成列名（排除元数据字段）
            metadata_fields = {"sample_time", "test_date", "created_at", "updated_at", "verified_at", "verified_by"}
            field_cols = []
            seen = set()
            for assessment, _, _ in results:
                if assessment.extracted_data:
                    for f in assessment.extracted_data.keys():
                        if f not in metadata_fields:
                            col_name = f"{assessment.assessment_type}_{f}"
                            if col_name not in seen:
                                field_cols.append(col_name)
                                seen.add(col_name)
            field_cols = sorted(field_cols)

            all_cols = ["样本编号", "访视名称", "访视日期", "检测类型", "是否审核", "采样时间", "创建时间"] + field_cols

            wb = Workbook()
            ws = wb.active
            ws.title = "Assessments"
            ws.append(all_cols)

            for row in wide_data:
                ws.append([row.get(col, "") for col in all_cols])

            # 调整列宽
            for col in range(1, len(all_cols) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=assessments_wide_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )


@router.get("/{assessment_id}", response_model=AssessmentDataResponse)
async def get_assessment(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """获取检查数据详情"""
    assessment = db.query(AssessmentData).filter(AssessmentData.id == assessment_id).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="检查数据不存在")
    return assessment


@router.put("/{assessment_id}", response_model=AssessmentDataResponse)
async def update_assessment(
    assessment_id: int,
    assessment_update: AssessmentDataUpdate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """更新检查数据（用于人工审核后确认）"""
    db_assessment = db.query(AssessmentData).filter(AssessmentData.id == assessment_id).first()
    if not db_assessment:
        raise HTTPException(status_code=404, detail="检查数据不存在")

    update_data = assessment_update.model_dump(exclude_unset=True)

    # 如果是确认审核，设置审核时间和审核人
    if assessment_update.is_verified and not db_assessment.is_verified:
        update_data["verified_at"] = datetime.utcnow()
        update_data["verified_by"] = current_user.id

    for key, value in update_data.items():
        setattr(db_assessment, key, value)

    db.commit()
    db.refresh(db_assessment)
    return db_assessment


@router.delete("/{assessment_id}")
async def delete_assessment(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """删除检查数据"""
    db_assessment = db.query(AssessmentData).filter(AssessmentData.id == assessment_id).first()
    if not db_assessment:
        raise HTTPException(status_code=404, detail="检查数据不存在")

    db.delete(db_assessment)
    db.commit()
    return {"message": "检查数据已删除"}


@router.get("/visit/{visit_id}/summary")
async def get_visit_summary(
    visit_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """获取随访数据汇总（用于前端审核界面）"""
    visit = db.query(Visit).filter(Visit.id == visit_id).first()
    if not visit:
        raise HTTPException(status_code=404, detail="随访记录不存在")

    subject = db.query(Subject).filter(Subject.id == visit.subject_id).first()

    assessments = db.query(AssessmentData).filter(AssessmentData.visit_id == visit_id).all()

    # 获取待审核的文件
    from app.models.tables import RawFile
    pending_files = db.query(RawFile).filter(
        RawFile.visit_id == visit_id,
        RawFile.status == 'pending_review'
    ).all()

    return {
        "visit": {
            "id": visit.id,
            "visit_name": visit.visit_name,
            "visit_date": visit.visit_date,
        },
        "subject": {
            "id": subject.id,
            "subject_code": subject.subject_code,
            "name_pinyin": subject.name_pinyin,
            "gender": subject.gender,
        },
        "assessments": [
            {
                "id": a.id,
                "assessment_type": a.assessment_type,
                "extracted_data": a.extracted_data,
                "is_verified": a.is_verified,
                "file_id": a.file_id,
            }
            for a in assessments
        ],
        "pending_files": [
            {
                "id": f.id,
                "file_type": f.file_type,
                "original_filename": f.original_filename,
                "ai_extracted_data": f.ai_extracted_data,
                "status": f.status,
                "uploaded_at": f.uploaded_at.isoformat() if f.uploaded_at else None,
            }
            for f in pending_files
            if f.ai_extracted_data  # 只返回有 AI 提取数据的文件
        ]
    }


@router.get("/{assessment_id}/download")
async def download_assessment(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """下载单个评估数据"""
    assessment = db.query(AssessmentData).filter(AssessmentData.id == assessment_id).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="评估数据不存在")

    data = {
        "id": assessment.id,
        "assessment_type": assessment.assessment_type,
        "extracted_data": assessment.extracted_data,
        "is_verified": assessment.is_verified,
        "verified_at": assessment.verified_at.isoformat() if assessment.verified_at else None,
        "created_at": assessment.created_at.isoformat() if assessment.created_at else None,
        "updated_at": assessment.updated_at.isoformat() if assessment.updated_at else None,
    }

    return Response(
        content=json.dumps(data, indent=2, ensure_ascii=False),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=assessment_{assessment_id}.json"}
    )
